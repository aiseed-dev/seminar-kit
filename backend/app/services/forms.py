"""申込様式 xlsx の生成(神Excel を機械可読に設計する)。

様式 xlsx が唯一のフォーム定義。名前付きセルの一覧(NAMES)はここが正で、
読み取り(parse.py)も同じ定義名を参照する。セル座標はこのモジュールに閉じ、
他の場所でハードコードしない。
Excel の定義名にはハイフンが使えないため、定義名のみアンダースコアを使う。

様式は公開サイトに置かない(受領メール・印刷物・ブラウザ記入セッションで
個別に渡す)。secret を渡すと発行キー(HMAC)を埋め込み、parse が検証する
——申込者にパスワードを持たせない代わりに、発行済み様式の所持を資格にする。
"""

import hashlib
import hmac as hmaclib
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.models.course import Course

FORM_VER = 1
SHEET = "申込書"
JST = ZoneInfo("Asia/Tokyo")

# 参加場所の表示ラベル(申込者が見る値)⇔ DB の location 値
LOC_VENUE = "会場"
LOC_ONLINE = "オンライン"
LOC_SATELLITE = "サテライト"
LOC_JA = {"venue": LOC_VENUE, "online": LOC_ONLINE, "satellite": LOC_SATELLITE}

# 企業・担当者欄: 定義名 → (行, ラベル)。値はいずれも C 列
_COMPANY_ROWS: dict[str, tuple[int, str]] = {
    "company_kana": (5, "フリガナ"),
    "company_name": (6, "企業・団体名"),
    "postal_code": (7, "郵便番号"),
    "address": (8, "所在地"),
    "tel": (9, "電話番号"),
    "fax": (10, "FAX(任意)"),
    "contact_kana": (11, "フリガナ"),
    "contact_name": (12, "ご担当者名"),
    "contact_email": (13, "メールアドレス"),
}

# 受講者欄(最大3名): ヘッダ行の下に1名1行
ATT_FIELDS: dict[str, tuple[str, str]] = {
    "name": ("B", "氏名"),
    "kana": ("C", "フリガナ"),
    "role": ("D", "所属・役職"),
    "email": ("E", "メールアドレス"),
    "loc": ("F", "参加場所"),
}
_ATT_HEADER_ROW = 15
_ATT_ROWS = (16, 17, 18)
_NOTE_ROW = 20

# 定義名 → セル座標(様式の機械可読部の全リスト)
NAMES: dict[str, str] = {
    "course_id": "H1",
    "form_ver": "H2",
    "form_key": "H3",  # 発行キー(HMAC。捏造様式の検出)
    **{name: f"C{row}" for name, (row, _) in _COMPANY_ROWS.items()},
    **{
        f"att{i}_{field}": f"{col}{_ATT_ROWS[i - 1]}"
        for i in (1, 2, 3)
        for field, (col, _) in ATT_FIELDS.items()
    },
}

REQUIRED = [n for n in _COMPANY_ROWS if n != "fax"]  # 企業欄の必須(FAXのみ任意)

# 送信用テキスト(メール本文貼り付け)の項目名 → 定義名。
# 様式内蔵マクロが生成し、parse.parse_text が同じ表で読む
TEXT_KEYS: dict[str, str] = {
    "course_id": "講座ID",
    "form_ver": "様式版",
    "form_key": "発行キー",
    "company_kana": "企業名フリガナ",
    "company_name": "企業名",
    "postal_code": "郵便番号",
    "address": "所在地",
    "tel": "電話番号",
    "fax": "FAX",
    "contact_kana": "担当者フリガナ",
    "contact_name": "担当者名",
    "contact_email": "メールアドレス",
    **{
        f"att{i}_{field}": f"受講者{i}{suffix}"
        for i in (1, 2, 3)
        for field, suffix in (
            ("name", ""),
            ("kana", "フリガナ"),
            ("role", "所属"),
            ("email", "メール"),
            ("loc", "参加場所"),
        )
    },
}
TEXT_SHEET = "送信用テキスト"

# 不備メッセージ用の日本語ラベル
LABELS: dict[str, str] = {
    **{name: label for name, (_, label) in _COMPANY_ROWS.items()},
    **{f"att{i}_{f}": label for i in (1, 2, 3) for f, (_, label) in ATT_FIELDS.items()},
}


def loc_labels(course: Course) -> list[str]:
    """講座が提供する参加場所のドロップダウン選択肢。"""
    labels = []
    if course.allow_venue:
        labels.append(LOC_VENUE)
    if course.allow_online:
        labels.append(LOC_ONLINE)
    if course.allow_satellite:
        note = (course.satellite_note or "").replace(",", "、").strip()
        labels.append(f"{LOC_SATELLITE}({note})" if note else LOC_SATELLITE)
    return labels


def label_to_loc(label: str) -> str | None:
    """表示ラベル → DB の location 値。判別できなければ None。"""
    s = label.strip()
    if s.startswith(LOC_SATELLITE):
        return "satellite"
    if s.startswith(LOC_ONLINE):
        return "online"
    if s.startswith(LOC_VENUE):
        return "venue"
    return None


def jst(dt: datetime) -> str:
    """JST の日本語表記(様式・静的サイトで共用)。"""
    d = dt.astimezone(JST)
    return f"{d.year}年{d.month}月{d.day}日 {d:%H:%M}"


def form_key(course_id, secret: str) -> str:
    """様式の発行キー(講座ID+様式版に対する HMAC。保存不要で検証できる)。"""
    msg = f"{course_id}:{FORM_VER}".encode()
    return hmaclib.new(secret.encode(), msg, hashlib.sha256).hexdigest()[:16]


_FILL_INPUT = PatternFill("solid", fgColor="FFFDE7")  # 記入欄(薄い黄色)
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _draw(ws: Worksheet, course: Course, submit_addr: str) -> None:
    """レイアウト(ラベル・見出し・案内文)を描く。申込書・記入例で共通。"""
    ws["A1"] = "受講申込書"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = "講座名"
    ws["B2"] = course.title
    ws["B2"].font = Font(bold=True)
    ws["A3"] = "開催日"
    ws["B3"] = jst(course.starts_at)
    ws["A4"] = "申込期限"
    ws["B4"] = jst(course.apply_deadline)

    for row, label in _COMPANY_ROWS.values():
        ws[f"A{row}"] = label
        cell = ws[f"C{row}"]
        cell.fill = _FILL_INPUT
        cell.border = _BORDER

    ws[f"A{_ATT_HEADER_ROW}"] = "受講者(最大3名)"
    ws[f"A{_ATT_HEADER_ROW}"].font = Font(bold=True)
    for col, label in ATT_FIELDS.values():
        head = ws[f"{col}{_ATT_HEADER_ROW}"]
        head.value = label
        head.border = _BORDER
        head.alignment = Alignment(horizontal="center")
        for row in _ATT_ROWS:
            cell = ws[f"{col}{row}"]
            cell.fill = _FILL_INPUT
            cell.border = _BORDER

    ws[f"A{_NOTE_ROW}"] = (
        f"記入後、このファイルをメールに添付して {submit_addr} へお送りください。"
        "印刷して FAX でもお申し込みいただけます。"
    )

    ws.column_dimensions["A"].width = 16
    for col in "BCDE":
        ws.column_dimensions[col].width = 22
    ws.column_dimensions["F"].width = 18
    ws.print_area = f"A1:F{_NOTE_ROW}"
    ws.page_setup.fitToWidth = 1


def build(course: Course, submit_addr: str, secret: str = "") -> Workbook:
    """講座ごとの申込様式を組み立てる(1枚目=申込書、2枚目=記入例)。

    secret を渡すと発行キーを埋め込む(運用では必須。parse が検証する)。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    _draw(ws, course, submit_addr)

    # 機械可読メタ(印刷範囲外の列に置き、非表示にする)
    ws[NAMES["course_id"]] = str(course.id)
    ws[NAMES["form_ver"]] = FORM_VER
    if secret:
        ws[NAMES["form_key"]] = form_key(course.id, secret)
    ws.column_dimensions["H"].hidden = True

    for name, coord in NAMES.items():
        ref = f"{quote_sheetname(SHEET)}!${coord[0]}${coord[1:]}"
        wb.defined_names[name] = DefinedName(name, attr_text=ref)

    # 参加場所はドロップダウン(講座が提供する場所のみ)
    labels = loc_labels(course)
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(labels) + '"',
        allow_blank=True,
        showErrorMessage=True,
        error="リストから選んでください",
    )
    ws.add_data_validation(dv)
    for i in (1, 2, 3):
        dv.add(ws[NAMES[f"att{i}_loc"]])

    # 入力セル以外はシート保護
    for name, coord in NAMES.items():
        if name not in ("course_id", "form_ver", "form_key"):
            ws[coord].protection = Protection(locked=False)
    ws.protection.sheet = True

    _example(wb, course, submit_addr, labels)
    _text_sheet(wb, submit_addr)
    return wb


TEXT_ROW0 = 6  # 送信用テキストの本文が始まる行(A列。数式で自動生成)


def _text_sheet(wb: Workbook, submit_addr: str) -> None:
    """「送信用テキスト」シート。

    A列: 数式で本文を1行1項目で自動生成(Excel・LibreOffice・OnlyOffice の
    どれでも動く)。行を範囲選択してコピー→メール本文に貼るだけ。
    C列: 内蔵マクロ(OnlyOffice 専用)が一括コピー用に1セルへまとめる補助。
    """
    ws = wb.create_sheet(TEXT_SHEET)
    ws["A1"] = "送信用テキスト(メール本文に貼り付けて送る)"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = (
        "おすすめ: 無料の OnlyOffice で開き、マクロ「送信用テキスト」を"
        "実行すると、未記入のチェック付きで C 列に本文がまとまります。"
        f"コピーして {submit_addr} へメール本文として送信してください。"
    )
    ws["A3"] = (
        "Excel 等では下の A 列に本文が自動で出ます(チェックなし)。"
        f"{TEXT_ROW0}行目から下を選択してコピーしてください。"
        "このファイルをそのまま添付して送っていただいても構いません。"
    )
    ws[f"A{TEXT_ROW0 - 1}"] = "↓ ここから下をコピー"
    ws[f"A{TEXT_ROW0 - 1}"].font = Font(bold=True)
    for i, (name, coord) in enumerate(NAMES.items()):
        ref = f"'{SHEET}'!{coord}"
        key = TEXT_KEYS[name]
        ws.cell(
            row=TEXT_ROW0 + i,
            column=1,
            value=f'=IF({ref}="","","{key}: "&{ref})',
        )
    ws.column_dimensions["A"].width = 90


def macro_js() -> str:
    """様式内蔵マクロ(OnlyOffice JavaScript)を様式定義から生成する。

    本文の生成自体は A 列の数式が担う(全表計算ソフト共通)。マクロは
    OnlyOffice 利用者向けの補助で、C 列の1セルに本文をまとめて
    一括コピーしやすくする。
    出力: `python -m app.services.forms > deploy/form-macro.js`
    座標・項目名は NAMES / TEXT_KEYS が正なので、様式を変えたら再生成する。
    導入時に OnlyOffice で様式に組み込む(xlsx への自動埋め込みは
    実機検証後の課題)。保守は North Mini Code の演習題材(docs/05)。
    """
    pairs = ",\n    ".join(
        f'["{TEXT_KEYS[name]}", "{coord}"]' for name, coord in NAMES.items()
    )
    # 未記入チェックの定義(サーバー側 parse と同じ REQUIRED / LABELS から生成)
    required = ",\n    ".join(
        f'["{LABELS[name]}", "{NAMES[name]}"]' for name in REQUIRED
    )
    def att_js(i: int) -> str:
        fields = ", ".join(
            f'["{label}", "{NAMES[f"att{i}_{field}"]}"]'
            for field, (_, label) in ATT_FIELDS.items()
        )
        return f"{{ n: {i}, f: [{fields}] }}"

    att_rows = ",\n    ".join(att_js(i) for i in (1, 2, 3))
    head = (
        f"// 申込様式マクロ: 未記入チェックのうえ、送信用テキストを"
        f"「{TEXT_SHEET}」シートに書き出す\n"
    )
    return (
        head
        + f"""// OnlyOffice(Desktop / Docs)専用。このファイルは自動生成
// (python -m app.services.forms)——手で直さず、様式(forms.py)を直して再生成する
// チェック定義はサーバー側の検証(parse)と同じ forms.py から生成される
(function () {{
  var src = Api.GetSheet("{SHEET}");
  var out = Api.GetSheet("{TEXT_SHEET}");
  function val(coord) {{
    var v = src.GetRange(coord).GetValue();
    return v === null ? "" : String(v).trim();
  }}

  // 1) 未記入チェック(正の検証はサーバー側。ここは送信前の親切)
  var issues = [];
  var required = [
    {required}
  ];
  required.forEach(function (kv) {{
    if (val(kv[1]) === "") issues.push("「" + kv[0] + "」が未記入です");
  }});
  var atts = [
    {att_rows}
  ];
  var entrants = 0;
  atts.forEach(function (a) {{
    var vals = a.f.map(function (kv) {{ return val(kv[1]); }});
    var any = vals.some(function (v) {{ return v !== ""; }});
    if (!any) return;
    var missing = [];
    a.f.forEach(function (kv, i) {{ if (vals[i] === "") missing.push(kv[0]); }});
    if (missing.length) {{
      issues.push("受講者" + a.n + "人目の「" + missing.join("・") + "」が未記入です");
    }} else {{
      entrants++;
    }}
  }});
  if (entrants === 0 && issues.length === 0) {{
    issues.push("受講者が1名も記入されていません");
  }}
  if (issues.length) {{
    out.GetRange("C{TEXT_ROW0 - 1}")
      .SetValue("【未記入があります。本文は作成されませんでした】");
    out.GetRange("C{TEXT_ROW0}").SetValue(issues.join("\\n"));
    out.SetActive();
    return;
  }}

  // 2) 送信用テキストの生成(1セルに一括コピー用)
  var map = [
    {pairs}
  ];
  var lines = [];
  map.forEach(function (kv) {{
    var v = val(kv[1]);
    if (v !== "") lines.push(kv[0] + ": " + v);
  }});
  out.GetRange("C{TEXT_ROW0 - 1}").SetValue("一括コピー用(マクロ出力)");
  out.GetRange("C{TEXT_ROW0}").SetValue(lines.join("\\n"));
  out.SetActive();
}})();
"""
    )


if __name__ == "__main__":
    print(macro_js())


_SAMPLE_COMPANY = {
    "company_kana": "カブシキガイシャ マルマルセイサクショ",
    "company_name": "株式会社 ○○製作所",
    "postal_code": "770-0000",
    "address": "徳島県徳島市○○町1-2-3",
    "tel": "088-000-0000",
    "fax": "088-000-0001",
    "contact_kana": "ヤマダ ハナコ",
    "contact_name": "山田 花子",
    "contact_email": "hanako@example.co.jp",
}


def _example(wb: Workbook, course: Course, submit_addr: str, labels: list[str]) -> None:
    """記入例シート(2枚目)。迷わせないための見本で、機械読み取りはしない。"""
    ws = wb.create_sheet("記入例")
    _draw(ws, course, submit_addr)
    for name, value in _SAMPLE_COMPANY.items():
        row, _ = _COMPANY_ROWS[name]
        ws[f"C{row}"] = value
    sample_att = {
        "name": "山田 太郎",
        "kana": "ヤマダ タロウ",
        "role": "製造部 課長",
        "email": "taro@example.co.jp",
        "loc": labels[0] if labels else LOC_VENUE,
    }
    for field, (col, _) in ATT_FIELDS.items():
        ws[f"{col}{_ATT_ROWS[0]}"] = sample_att[field]
    ws.protection.sheet = True
