"""当日名簿の xlsx 出力(openpyxl 直接生成)。

印刷して受付で紙チェックする運用(タップ入力はしない)。
Microsoft 365 不要——OnlyOffice / LibreOffice で開いて印刷できる。
"""

from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.models import Course
from app.services.forms import LOC_JA, jst

_LOC_ORDER = {"venue": 0, "satellite": 1, "online": 2}


@dataclass(frozen=True)
class Row:
    """名簿1行(queries が DB から組み立てる)。"""

    application_no: str
    company_name: str
    name: str
    kana: str
    location: str


_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FILL = PatternFill("solid", fgColor="EEF3F8")

_HEADERS = ("No", "申込番号", "企業・団体名", "氏名", "フリガナ", "参加場所", "出席")


def build(course: Course, rows: list[Row]) -> Workbook:
    """当日名簿を組み立てる(会場→サテライト→オンライン、フリガナ順)。"""
    rows = sorted(rows, key=lambda r: (_LOC_ORDER.get(r.location, 9), r.kana))

    wb = Workbook()
    ws = wb.active
    ws.title = "当日名簿"

    ws["A1"] = f"当日名簿 {course.title}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"開催日時: {jst(course.starts_at)}"
    counts = {}
    for r in rows:
        counts[r.location] = counts.get(r.location, 0) + 1
    ws["A3"] = "参加内訳: " + " / ".join(
        f"{LOC_JA[loc]} {n}名"
        for loc, n in sorted(counts.items(), key=lambda x: _LOC_ORDER.get(x[0], 9))
    )

    head_row = 5
    for col, head in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=head_row, column=col, value=head)
        cell.font = Font(bold=True)
        cell.fill = _HEAD_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, start=1):
        values = (
            i,
            r.application_no,
            r.company_name,
            r.name,
            r.kana,
            LOC_JA.get(r.location, r.location),
            "",  # 出席チェック欄(紙でチェック)
        )
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=head_row + i, column=col, value=v)
            cell.border = _BORDER

    widths = (5, 12, 26, 16, 18, 12, 8)
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + col - 1)].width = w

    ws.print_title_rows = f"{head_row}:{head_row}"  # 2枚目以降もヘッダを繰り返す
    ws.print_area = f"A1:G{head_row + len(rows)}"
    ws.page_setup.fitToWidth = 1
    return wb
