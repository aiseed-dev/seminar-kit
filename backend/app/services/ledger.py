"""年度実績台帳の xlsx / CSV 出力(事業報告の基礎資料)。

年度は4月起点(1〜3月は前年度)。集計値の算出は office/queries が行い、
ここは表の組み立てに徹する。
"""

import csv
import io
from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from app.services.forms import JST, jst


@dataclass(frozen=True)
class Row:
    """台帳1行=講座1件(年度内)。"""

    title: str
    category: str
    starts_at: datetime
    status: str
    applications: int  # 申込数(受付済みの申込件数)
    attendees: int  # 確定数(受付済み申込の受講者数)
    attendance: int | None  # 出席数(開催後入力。未入力は None)


_STATUS_JA = {"open": "募集中", "closed": "締切", "finished": "開催済み"}
_HEADERS = ("講座名", "分類", "開催日", "状態", "申込数", "確定数", "出席数")


def fiscal_year(dt: datetime) -> int:
    """年度(4月起点)。1〜3月は前年度に属する。"""
    d = dt.astimezone(JST)
    return d.year if d.month >= 4 else d.year - 1


def _table(year: int, rows: list[Row]) -> list[tuple]:
    """ヘッダ+明細+合計の行列(xlsx と CSV で共用)。"""
    rows = sorted(rows, key=lambda r: r.starts_at)
    out: list[tuple] = [(f"{year}年度 実績台帳",), _HEADERS]
    for r in rows:
        out.append(
            (
                r.title,
                r.category,
                jst(r.starts_at),
                _STATUS_JA.get(r.status, r.status),
                r.applications,
                r.attendees,
                r.attendance if r.attendance is not None else "",
            )
        )
    out.append(
        (
            f"合計({len(rows)}講座)",
            "",
            "",
            "",
            sum(r.applications for r in rows),
            sum(r.attendees for r in rows),
            sum(r.attendance or 0 for r in rows),
        )
    )
    return out


def build(year: int, rows: list[Row]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年度"
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    table = _table(year, rows)
    for i, row in enumerate(table, start=1):
        for col, v in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=v)
            if i >= 2:
                cell.border = border
    ws["A1"].font = Font(size=14, bold=True)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEF3F8")
    ws.cell(row=len(table), column=1).font = Font(bold=True)
    widths = (36, 14, 20, 10, 8, 8, 8)
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + col - 1)].width = w
    return wb


def to_csv(year: int, rows: list[Row]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(_table(year, rows))
    return buf.getvalue()
