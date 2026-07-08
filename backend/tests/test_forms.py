"""様式の生成(forms)と読み取り(parse)のラウンドトリップ検査。"""

import io
import uuid
from datetime import datetime

import pytest
from app.models import Course
from app.services import forms, parse
from openpyxl import Workbook, load_workbook

JST = forms.JST
SUBMIT = "moshikomi@example.jp"


def make_course(**over) -> Course:
    base = dict(
        id=uuid.uuid4(),
        title="DX入門セミナー",
        starts_at=datetime(2026, 9, 1, 13, 30, tzinfo=JST),
        apply_deadline=datetime(2026, 8, 25, 17, 0, tzinfo=JST),
        allow_venue=True,
        allow_online=True,
        allow_satellite=False,
        satellite_note=None,
    )
    base.update(over)
    return Course(**base)


def build_wb(course: Course):
    buf = io.BytesIO()
    forms.build(course, SUBMIT).save(buf)
    buf.seek(0)
    return load_workbook(buf)


def set_named(wb, name: str, value) -> None:
    ((title, coord),) = wb.defined_names[name].destinations
    wb[title][coord] = value


def dump(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


COMPANY = {
    "company_kana": "カブシキガイシャテスト",
    "company_name": "株式会社テスト",
    "postal_code": "770-8570",
    "address": "徳島県徳島市万代町1-1",
    "tel": "088-621-2323",
    "contact_kana": "トクシマ タロウ",
    "contact_name": "徳島 太郎",
    "contact_email": "taro@test.example.jp",
}


def fill_company(wb) -> None:
    for name, value in COMPANY.items():
        set_named(wb, name, value)


def fill_att(wb, i: int, loc: str = "会場") -> None:
    set_named(wb, f"att{i}_name", f"受講 {i}郎")
    set_named(wb, f"att{i}_kana", f"ジュコウ {i}ロウ")
    set_named(wb, f"att{i}_role", "総務部")
    set_named(wb, f"att{i}_email", f"user{i}@test.example.jp")
    set_named(wb, f"att{i}_loc", loc)


def test_roundtrip():
    course = make_course()
    wb = build_wb(course)
    fill_company(wb)
    fill_att(wb, 1, "会場")
    got = parse.parse(dump(wb))

    assert got.course_id == course.id
    assert got.form_ver == forms.FORM_VER
    assert got.company_name == COMPANY["company_name"]
    assert got.contact_email == COMPANY["contact_email"]
    assert got.fax is None  # 任意欄は未記入で通る
    (entrant,) = got.entrants
    assert entrant.name == "受講 1郎"
    assert entrant.loc == "venue"


def test_roundtrip_three_entrants():
    wb = build_wb(make_course())
    fill_company(wb)
    fill_att(wb, 1, "会場")
    fill_att(wb, 2, "オンライン")
    fill_att(wb, 3, "会場")
    got = parse.parse(dump(wb))
    assert [e.loc for e in got.entrants] == ["venue", "online", "venue"]


def test_satellite_label_parses():
    course = make_course(allow_satellite=True, satellite_note="toku-Noix")
    labels = forms.loc_labels(course)
    assert labels == ["会場", "オンライン", "サテライト(toku-Noix)"]
    wb = build_wb(course)
    fill_company(wb)
    fill_att(wb, 1, "サテライト(toku-Noix)")
    got = parse.parse(dump(wb))
    assert got.entrants[0].loc == "satellite"


def test_missing_company_field():
    wb = build_wb(make_course())
    fill_company(wb)
    set_named(wb, "company_name", None)
    fill_att(wb, 1)
    with pytest.raises(parse.Invalid) as e:
        parse.parse(dump(wb))
    assert any("企業・団体名" in s for s in e.value.issues)


def test_no_entrants():
    wb = build_wb(make_course())
    fill_company(wb)
    with pytest.raises(parse.Invalid) as e:
        parse.parse(dump(wb))
    assert any("受講者が1名も" in s for s in e.value.issues)


def test_partial_entrant_row():
    wb = build_wb(make_course())
    fill_company(wb)
    fill_att(wb, 1)
    set_named(wb, "att2_name", "名前だけ")  # 2人目は氏名のみ記入
    with pytest.raises(parse.Invalid) as e:
        parse.parse(dump(wb))
    assert any("受講者2人目" in s for s in e.value.issues)


def test_bad_location():
    wb = build_wb(make_course())
    fill_company(wb)
    fill_att(wb, 1, "自宅")
    with pytest.raises(parse.Invalid) as e:
        parse.parse(dump(wb))
    assert any("参加場所" in s for s in e.value.issues)


def test_wrong_version():
    wb = build_wb(make_course())
    fill_company(wb)
    fill_att(wb, 1)
    set_named(wb, "form_ver", 99)
    with pytest.raises(parse.Invalid) as e:
        parse.parse(dump(wb))
    assert any("版" in s for s in e.value.issues)


def test_not_a_form():
    blank = Workbook()
    with pytest.raises(parse.Invalid) as e:
        parse.parse(dump(blank))
    assert any("様式のファイルではない" in s for s in e.value.issues)


def test_form_is_protected_and_machine_readable():
    wb = build_wb(make_course())
    ws = wb[forms.SHEET]
    assert ws.protection.sheet  # シート保護
    assert ws.column_dimensions["H"].hidden  # メタ列は非表示
    # 入力セルは保護解除、メタセルは保護のまま
    assert ws[forms.NAMES["company_name"]].protection.locked is False
    assert ws[forms.NAMES["course_id"]].protection.locked is True
    assert ws[forms.NAMES["form_key"]].protection.locked is True
    # 参加場所のドロップダウン(入力規則)がある
    assert len(ws.data_validations.dataValidation) == 1
    # 記入例シートが付いている
    assert "記入例" in wb.sheetnames


def test_all_named_cells_registered():
    wb = build_wb(make_course())
    for name in forms.NAMES:
        assert name in wb.defined_names


# ---- 発行キー(パスワードの代わりに、発行済み様式の所持を資格にする) ----

SECRET = "form-secret-for-test"


def build_wb_with_key(course, secret=SECRET):
    buf = io.BytesIO()
    forms.build(course, SUBMIT, secret=secret).save(buf)
    buf.seek(0)
    return load_workbook(buf)


def test_form_key_roundtrip():
    course = make_course()
    wb = build_wb_with_key(course)
    fill_company(wb)
    fill_att(wb, 1)
    got = parse.parse(dump(wb), secret=SECRET)
    assert got.course_id == course.id


def test_forged_form_rejected():
    # 発行キーなし(捏造相当)の様式は、検証ありの読み取りで弾く
    wb = build_wb(make_course())
    fill_company(wb)
    fill_att(wb, 1)
    with pytest.raises(parse.Invalid) as e:
        parse.parse(dump(wb), secret=SECRET)
    assert any("発行元" in s for s in e.value.issues)


def test_wrong_secret_rejected():
    wb = build_wb_with_key(make_course(), secret="other-secret")
    fill_company(wb)
    fill_att(wb, 1)
    with pytest.raises(parse.Invalid):
        parse.parse(dump(wb), secret=SECRET)


def test_key_not_checked_when_secret_off():
    # secret 未設定(開発時)は検証しない
    wb = build_wb(make_course())
    fill_company(wb)
    fill_att(wb, 1)
    parse.parse(dump(wb))  # 例外なし


def test_form_key_is_stable_per_course():
    course = make_course()
    assert forms.form_key(course.id, SECRET) == forms.form_key(course.id, SECRET)
    assert forms.form_key(course.id, SECRET) != forms.form_key(course.id, "x")


# ---- 送信用テキスト(本文貼り付け。様式内蔵マクロが生成する形式) ----

TEXT_COMPANY = {
    "企業名フリガナ": "カブシキガイシャテスト",
    "企業名": "株式会社テスト",
    "郵便番号": "770-8570",
    "所在地": "徳島県徳島市万代町1-1",
    "電話番号": "088-621-2323",
    "担当者フリガナ": "トクシマ タロウ",
    "担当者名": "徳島 太郎",
    "メールアドレス": "taro@test.example.jp",
}


def text_of(course, secret=None, drop=(), extra=()):
    vals = {
        "講座ID": str(course.id),
        "様式版": str(forms.FORM_VER),
        **TEXT_COMPANY,
        "受講者1": "受講 1郎",
        "受講者1フリガナ": "ジュコウ 1ロウ",
        "受講者1所属": "総務部",
        "受講者1メール": "user1@test.example.jp",
        "受講者1参加場所": "会場",
    }
    if secret:
        vals["発行キー"] = forms.form_key(course.id, secret)
    for k in drop:
        vals.pop(k)
    lines = [f"{k}: {v}" for k, v in vals.items()]
    lines.extend(extra)
    return "\n".join(lines)


def test_parse_text_roundtrip():
    course = make_course()
    got = parse.parse_text(text_of(course))
    assert got.course_id == course.id
    assert got.company_name == "株式会社テスト"
    (e,) = got.entrants
    assert (e.name, e.loc) == ("受講 1郎", "venue")


def test_read_text_fields_for_prefill():
    """代行入力の前埋め用: 検証なしで内部名→値を返す(部分でも読める)。"""
    course = make_course()
    vals = parse.read_text_fields(text_of(course))
    assert vals["company_name"] == "株式会社テスト"
    assert vals["contact_email"] == "taro@test.example.jp"
    assert vals["att1_name"] == "受講 1郎"
    assert vals["att1_loc"] == "会場"  # ラベルのまま(コード変換は呼び出し側)


def test_read_text_fields_partial_no_course():
    """講座ID が無い部分的な貼り付けでも、読める項目だけ返す。"""
    vals = parse.read_text_fields("企業名: 株式会社テスト\n担当者名: 徳島 太郎")
    assert vals == {"company_name": "株式会社テスト", "contact_name": "徳島 太郎"}


def test_read_text_fields_empty():
    assert parse.read_text_fields("ただの文章です。項目はありません。") == {}


def test_parse_text_with_key():
    course = make_course()
    got = parse.parse_text(text_of(course, secret=SECRET), secret=SECRET)
    assert got.course_id == course.id


def test_parse_text_forged_rejected():
    course = make_course()
    with pytest.raises(parse.Invalid) as e:
        parse.parse_text(text_of(course), secret=SECRET)  # キーなし
    assert any("発行元" in s for s in e.value.issues)


def test_parse_text_tolerates_quoting_and_zenkaku():
    """返信の引用(>)・全角コロン・余計な文にも耐える。"""
    course = make_course()
    body = (
        "お世話になります。\n"
        + "\n".join(
            "> " + line.replace(": ", ": ", 1) for line in text_of(course).splitlines()
        )
        + "\nよろしくお願いします。"
    )
    got = parse.parse_text(body)
    assert got.company_name == "株式会社テスト"


def test_parse_text_not_a_form():
    assert parse.parse_text("こんにちは。DX入門セミナーに申し込みます。") is None


def test_parse_text_missing_required():
    course = make_course()
    with pytest.raises(parse.Invalid) as e:
        parse.parse_text(text_of(course, drop=("企業名",)))
    assert any("企業・団体名" in s for s in e.value.issues)


def test_text_keys_cover_all_names():
    assert set(forms.TEXT_KEYS) == set(forms.NAMES)


def test_macro_js_generated_from_definition():
    js = forms.macro_js()
    for name, coord in forms.NAMES.items():
        assert f'"{coord}"' in js
        assert f'"{forms.TEXT_KEYS[name]}"' in js
    assert forms.SHEET in js and forms.TEXT_SHEET in js
    # 未記入チェックもサーバー側と同じ定義から生成される
    for name in forms.REQUIRED:
        assert f'"{forms.LABELS[name]}"' in js
    assert "未記入" in js
    assert "受講者が1名も" in js


def test_text_sheet_has_formulas():
    """送信用テキストは数式で自動生成(Excel でも本文貼り付けができる)。"""
    wb = build_wb(make_course())
    ws = wb[forms.TEXT_SHEET]
    for i, name in enumerate(forms.NAMES):
        formula = ws.cell(row=forms.TEXT_ROW0 + i, column=1).value
        assert formula.startswith("=IF(")
        assert forms.TEXT_KEYS[name] in formula
        assert forms.NAMES[name] in formula  # 参照座標


def test_parse_text_single_cell_paste():
    """マクロの1セル出力を貼ると引用符で囲まれることがある——それにも耐える。"""
    course = make_course()
    body = '"' + text_of(course) + '"'
    got = parse.parse_text(body)
    assert got.company_name == "株式会社テスト"
    assert got.entrants[0].loc == "venue"  # 末尾の引用符が付いても読める
