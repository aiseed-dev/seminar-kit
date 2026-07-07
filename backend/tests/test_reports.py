"""帳票(当日名簿・年度実績台帳)の検査(DB 不要)。"""

import io
import uuid
from datetime import datetime

from app.models import Course
from app.services import ledger, roster
from app.services.forms import JST
from openpyxl import load_workbook

STARTS = datetime(2026, 9, 1, 13, 30, tzinfo=JST)


def load(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf)


# ---- 当日名簿 ----


def make_rows():
    return [
        roster.Row("2026-00002", "B社", "後井 次郎", "ゴイ ジロウ", "online"),
        roster.Row("2026-00001", "A社", "前田 太郎", "マエダ タロウ", "venue"),
        roster.Row("2026-00003", "C社", "阿部 花子", "アベ ハナコ", "venue"),
    ]


def test_roster_build():
    course = Course(id=uuid.uuid4(), title="DX入門セミナー", starts_at=STARTS)
    wb = load(roster.build(course, make_rows()))
    ws = wb["当日名簿"]
    assert "DX入門セミナー" in ws["A1"].value
    assert "会場 2名" in ws["A3"].value and "オンライン 1名" in ws["A3"].value
    # 並び: 会場(フリガナ順: アベ→マエダ)→ オンライン
    names = [ws.cell(row=r, column=4).value for r in (6, 7, 8)]
    assert names == ["阿部 花子", "前田 太郎", "後井 次郎"]
    # 出席チェック欄は空
    assert all(not ws.cell(row=r, column=7).value for r in (6, 7, 8))
    # 2枚目以降のヘッダ繰り返し
    assert ws.print_title_rows == "$5:$5"


def test_roster_empty():
    course = Course(id=uuid.uuid4(), title="空講座", starts_at=STARTS)
    ws = load(roster.build(course, []))["当日名簿"]
    assert ws.cell(row=5, column=1).value == "No"


# ---- 年度実績台帳 ----


def test_fiscal_year():
    assert ledger.fiscal_year(datetime(2026, 4, 1, tzinfo=JST)) == 2026
    assert ledger.fiscal_year(datetime(2027, 3, 31, tzinfo=JST)) == 2026
    assert ledger.fiscal_year(datetime(2027, 4, 1, tzinfo=JST)) == 2027


def make_ledger_rows():
    return [
        ledger.Row("講座A", "人材育成", STARTS, "finished", 10, 15, 14),
        ledger.Row("講座B", "DX", STARTS, "closed", 5, 7, None),
    ]


def test_ledger_build():
    ws = load(ledger.build(2026, make_ledger_rows())).active
    assert ws["A1"].value == "2026年度 実績台帳"
    assert ws["A2"].value == "講座名"
    # 明細
    assert ws["A3"].value == "講座A"
    assert ws["G3"].value == 14
    assert not ws["G4"].value  # 出席未入力は空欄
    # 合計行
    assert ws["A5"].value == "合計(2講座)"
    assert (ws["E5"].value, ws["F5"].value, ws["G5"].value) == (15, 22, 14)


def test_ledger_csv():
    text = ledger.to_csv(2026, make_ledger_rows())
    lines = text.strip().splitlines()
    assert lines[0].startswith("2026年度")
    assert "講座A" in text and "合計(2講座)" in lines[-1]
